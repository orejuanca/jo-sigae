#!/usr/bin/env python3
"""
Genera un Excel con todos los bindings del sistema jo-sigae organizados en columnas claras.
Columnas:
  A: Grupo / Categoria
  B: Binding (value) - el string que se pone en dataBinding
  C: Descripcion (label) - lo que se ve en el editor
  D: Fuente de Datos - de donde viene el dato (rawData, student, doc, etc.)
  E: Clave rawData - la clave en rawDataMap (si aplica)
  F: Campo DisplayData - el campo en DisplayData (si aplica)
  G: Notas adicionales
"""

import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# =============================================
# HOJA 1: PLAN VIGENTE
# =============================================
ws_v = wb.active
ws_v.title = "Plan Vigente"

# Colores
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
GROUP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GROUP_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
SUBGROUP_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
RAW_FONT = Font(name="Consolas", size=10, color="006100")  # green for rawData
STRUCT_FONT = Font(name="Consolas", size=10, color="C00000")  # red for structured
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
WRAP = Alignment(wrap_text=True, vertical='top')

# Headers Vigente
headers_v = [
    "Grupo", "Binding (dataBinding)", "Descripcion (Editor)",
    "Fuente de Datos", "Clave rawDataMap", "Campo en DisplayData", "Notas"
]
for col, h in enumerate(headers_v, 1):
    cell = ws_v.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    cell.border = THIN_BORDER

# Widths
widths_v = [25, 38, 45, 22, 30, 35, 40]
for i, w in enumerate(widths_v, 1):
    ws_v.column_dimensions[get_column_letter(i)].width = w
ws_v.row_dimensions[1].height = 30

row = 2

def add_group_row(ws, r, group_name):
    for col in range(1, 8):
        c = ws.cell(row=r, column=col)
        c.fill = GROUP_FILL
        c.border = THIN_BORDER
    ws.cell(row=r, column=1, value=group_name).font = GROUP_FONT

def add_data_row(ws, r, binding, label, fuente, raw_key, display_field, notas, is_raw=False):
    vals = ["", binding, label, fuente, raw_key, display_field, notas]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.border = THIN_BORDER
        c.alignment = WRAP
        if col == 2:
            c.font = RAW_FONT if is_raw else STRUCT_FONT
        else:
            c.font = DATA_FONT

# ----- ESTUDIANTE -----
add_group_row(ws_v, row, "ESTUDIANTE"); row += 1
est_rows = [
    ("student.cedula", "Cedula de Identidad", "DisplayData.estudiante.cedula", "", "estudiante.cedula", "Datos del estudiante"),
    ("student.apellidos", "Apellidos", "DisplayData.estudiante.apellidos", "", "estudiante.apellidos", ""),
    ("student.nombres", "Nombres", "DisplayData.estudiante.nombres", "", "estudiante.nombres", ""),
    ("student.fechaNacimiento", "Fecha de Nacimiento", "DisplayData.estudiante.fechaNacimiento", "", "estudiante.fechaNacimiento", ""),
    ("student.pais", "Pais de Nacimiento", "DisplayData.estudiante.pais", "", "estudiante.pais", ""),
    ("student.estado", "Estado de Nacimiento", "DisplayData.estudiante.estado", "", "estudiante.estado", ""),
    ("student.municipio", "Municipio de Nacimiento", "DisplayData.estudiante.municipio", "", "estudiante.municipio", ""),
    ("student.lugarFechaNac", "Lugar y Fecha de Nacimiento", "DisplayData.estudiante (combinado)", "", "estudiante (compuesto)", "Concatena lugar + fecha"),
]
for b, l, f, rk, df, n in est_rows:
    add_data_row(ws_v, row, b, l, f, rk, df, n, False); row += 1

# ----- INSTITUCION (ESCUELA) -----
add_group_row(ws_v, row, "INSTITUCION (ESCUELA)"); row += 1
school_rows = [
    ("school.codigo", "Codigo OD", "DisplayData.od", "", "od", "Codigo de la escuela"),
    ("school.denominacion", "Denominacion", "DisplayData.denominacion", "", "denominacion", "Nombre de la escuela"),
    ("school.direccion", "Direccion", "DisplayData.direccion", "", "direccion", ""),
    ("school.telefono", "Telefono", "DisplayData.telefono", "", "telefono", ""),
    ("school.municipio", "Municipio", "DisplayData.municipio", "", "municipio", ""),
    ("school.estado", "Estado", "DisplayData.estado", "", "estado", ""),
    ("school.cdcce", "CDCEE", "DisplayData.cdcce", "", "cdcce", "Codigo del ente descentralizado"),
]
for b, l, f, rk, df, n in school_rows:
    add_data_row(ws_v, row, b, l, f, rk, df, n, False); row += 1

# ----- DOCUMENTO -----
add_group_row(ws_v, row, "DOCUMENTO"); row += 1
doc_rows = [
    ("doc.planEstudio", "Plan de Estudio", "DisplayData.planEstudio", "", "planEstudio", "Ej: 'PLAN VIGENTE'"),
    ("doc.codigo", "Codigo Plan", "DisplayData.planCodigo", "", "planCodigo", "Codigo numerico del plan"),
    ("doc.lugar", "Lugar", "DisplayData.lugar", "", "lugar", "Lugar de expedicion"),
    ("doc.fechaExpedicion", "Fecha de Expedicion", "DisplayData.fechaExpedicion", "", "fechaExpedicion", ""),
    ("doc.observaciones", "Observaciones", "DisplayData.observaciones", "", "observaciones", ""),
    ("doc.promedioAcumulado", "Promedio Acumulado", "DisplayData.promedioAcumulado", "", "promedioAcumulado", ""),
    ("doc.acta", "Serial Titulo (Acta)", "DisplayData.acta", "", "acta", "Serial del titulo"),
    ("doc.actaFecha", "Fecha Emision Titulo", "DisplayData.actaFecha", "", "actaFecha", ""),
    ("doc.actaAnio", "Anio Egreso Titulo", "DisplayData.actaAnio", "", "actaAnio", ""),
    ("rawData.CERT.EXPEDICION", "Fecha Emision N.", "rawDataMap", "CERT.EXPEDICION", "", "Via rawData - Fecha emision notas"),
    ("rawData.TITULO.SERIAL", "Serial Titulo (raw)", "rawDataMap", "TITULO.SERIAL", "", "Via rawData"),
    ("rawData.TITULO.EXPEDICION", "Fecha Emision T. (raw)", "rawDataMap", "TITULO.EXPEDICION", "", "Via rawData"),
    ("rawData.TITULO.EGRESO", "Anio Egreso T. (raw)", "rawDataMap", "TITULO.EGRESO", "", "Via rawData"),
]
for b, l, f, rk, df, n in doc_rows:
    is_raw = b.startswith("rawData.")
    add_data_row(ws_v, row, b, l, f, rk, df, n, is_raw); row += 1

# ----- LITERALES FINALES -----
add_group_row(ws_v, row, "LITERALES FINALES"); row += 1
lit_rows = [
    ("doc.literalFinal.0", "Literal Final - 1er Anio", "DisplayData.literalesFinales[0]", "", "literalesFinales[0]", ""),
    ("doc.literalFinal.1", "Literal Final - 2do Anio", "DisplayData.literalesFinales[1]", "", "literalesFinales[1]", ""),
    ("doc.literalFinal.2", "Literal Final - 3er Anio", "DisplayData.literalesFinales[2]", "", "literalesFinales[2]", ""),
    ("doc.literalFinal.3", "Literal Final - 4to Anio", "DisplayData.literalesFinales[3]", "", "literalesFinales[3]", ""),
    ("doc.literalFinal.4", "Literal Final - 5to Anio", "DisplayData.literalesFinales[4]", "", "literalesFinales[4]", ""),
]
for b, l, f, rk, df, n in lit_rows:
    add_data_row(ws_v, row, b, l, f, rk, df, n, False); row += 1

# ----- OBSERVACIONES -----
add_group_row(ws_v, row, "OBSERVACIONES"); row += 1
obs_rows = [
    ("obsCert.0", "OBS.CERT - Linea 1", "DisplayData.observacionesLines[0]", "", "observacionesLines[0]", ""),
    ("obsCert.1", "OBS.CERT - Linea 2", "DisplayData.observacionesLines[1]", "", "observacionesLines[1]", ""),
    ("obsCert.2", "OBS.CERT - Linea 3", "DisplayData.observacionesLines[2]", "", "observacionesLines[2]", ""),
    ("obsCert.3", "OBS.CERT - Linea 4", "DisplayData.observacionesLines[3]", "", "observacionesLines[3]", ""),
    ("rawData.OBS.NOTAS.L1", "Obs. Notas - Linea 1", "rawDataMap", "OBS.NOTAS.L1", "", "Via rawData"),
    ("rawData.OBS.NOTAS.L2", "Obs. Notas - Linea 2", "rawDataMap", "OBS.NOTAS.L2", "", "Via rawData"),
    ("rawData.OBS.NOTAS.L3", "Obs. Notas - Linea 3", "rawDataMap", "OBS.NOTAS.L3", "", "Via rawData"),
    ("rawData.OBS.BOLETA.L1", "Obs. Boleta - Linea 1", "rawDataMap", "OBS.BOLETA.L1", "", "Via rawData"),
    ("rawData.OBS.BOLETA.L2", "Obs. Boleta - Linea 2", "rawDataMap", "OBS.BOLETA.L2", "", "Via rawData"),
    ("rawData.OBS.BOLETA.L3", "Obs. Boleta - Linea 3", "rawDataMap", "OBS.BOLETA.L3", "", "Via rawData"),
]
for b, l, f, rk, df, n in obs_rows:
    is_raw = b.startswith("rawData.")
    add_data_row(ws_v, row, b, l, f, rk, df, n, is_raw); row += 1

# ----- DIRECTOR -----
add_group_row(ws_v, row, "DIRECTOR"); row += 1
dir_rows = [
    ("director.nombre", "Nombre del Director", "rawDataMap.DIRECTOR.NOMBRE o DisplayData.director.apellidosNombres", "DIRECTOR.NOMBRE (si viene de dashboard)", "director.apellidosNombres", "Primero busca en rawDataMap, fallback a DisplayData"),
    ("director.cedula", "Cedula del Director", "rawDataMap.DIRECTOR.CEDULA o DisplayData.director.cedula", "DIRECTOR.CEDULA (si viene de dashboard)", "director.cedula", ""),
    ("expedicion.fecha", "Fecha de Expedicion", "rawDataMap.EXPEDICION.FECHA o DisplayData.fechaExpedicion", "EXPEDICION.FECHA", "fechaExpedicion", "Viene del dashboard (celda [3][25])"),
    ("expedicion.lugar", "Lugar de Expedicion", "rawDataMap.EXPEDICION.LUGAR o DisplayData.lugar", "EXPEDICION.LUGAR", "lugar", "Viene del dashboard (celda [3][33])"),
]
for b, l, f, rk, df, n in dir_rows:
    add_data_row(ws_v, row, b, l, f, rk, df, n, False); row += 1

# ----- DIRECTOR CDCEE -----
add_group_row(ws_v, row, "DIRECTOR CDCEE"); row += 1
cdcee_rows = [
    ("cdcee.nombre", "Nombre Director CDCEE", "DisplayData.directorCdcce.apellidosNombres", "", "directorCdcce.apellidosNombres", ""),
    ("cdcee.cedula", "Cedula Director CDCEE", "DisplayData.directorCdcce.cedula", "", "directorCdcce.cedula", ""),
]
for b, l, f, rk, df, n in cdcee_rows:
    add_data_row(ws_v, row, b, l, f, rk, df, n, False); row += 1

# ----- INSTITUCIONES EDUCATIVAS -----
add_group_row(ws_v, row, "INSTITUCIONES EDUCATIVAS"); row += 1
for i in range(5):
    inst_rows = [
        (f"inst.{i}.denominacion", f"Institucion {i+1} - Denominacion", f"DisplayData.instituciones[{i}].denominacion", "", f"instituciones[{i}].denominacion", ""),
        (f"inst.{i}.localidad", f"Institucion {i+1} - Localidad", f"DisplayData.instituciones[{i}].localidad", "", f"instituciones[{i}].localidad", ""),
        (f"inst.{i}.ef", f"Institucion {i+1} - E.F.", f"DisplayData.instituciones[{i}].ef", "", f"instituciones[{i}].ef", ""),
    ]
    for b, l, f, rk, df, n in inst_rows:
        add_data_row(ws_v, row, b, l, f, rk, df, n, False); row += 1

# ----- CALIFICACIONES (estructurado: calif.X.Y.Z) -----
YEAR_NAMES = ['Primer Anio', 'Segundo Anio', 'Tercer Anio', 'Cuarto Anio', 'Quinto Anio']
MATERIAS_VIGENTE = {
    1: ['Castellano', 'Ingles', 'Matematicas', 'Educacion Fisica', 'Arte y Patrimonio', 'Ciencias Naturales', 'Geografia, Historia y Ciudadania'],
    2: ['Castellano', 'Ingles', 'Matematicas', 'Educacion Fisica', 'Arte y Patrimonio', 'Ciencias Naturales', 'Geografia, Historia y Ciudadania'],
    3: ['Castellano', 'Ingles', 'Matematicas', 'Educacion Fisica', 'Fisica', 'Quimica', 'Biologia', 'Geografia, Historia y Ciudadania', 'Formacion Soberania Nacional'],
    4: ['Castellano', 'Ingles', 'Matematicas', 'Educacion Fisica', 'Fisica', 'Quimica', 'Biologia', 'Geografia, Historia y Ciudadania', 'Formacion Soberania Nacional'],
    5: ['Castellano', 'Ingles', 'Matematicas', 'Educacion Fisica', 'Fisica', 'Quimica', 'Biologia', 'Ciencias de la Tierra', 'Geografia, Historia y Ciudadania', 'Formacion Soberania Nacional'],
}

ABREV_MAP = {
    'Castellano': 'CA', 'Ingles': 'IN', 'Matematicas': 'MA', 'Educacion Fisica': 'EF',
    'Arte y Patrimonio': 'AP', 'Ciencias Naturales': 'CN',
    'Geografia, Historia y Ciudadania': 'GH', 'Formacion Soberania Nacional': 'FSN',
    'Fisica': 'FI', 'Quimica': 'QU', 'Biologia': 'BI', 'Ciencias de la Tierra': 'CT',
}

add_group_row(ws_v, row, "CALIFICACIONES (Estructurado: calif.anio.index.campo)"); row += 1

for y in range(1, 6):
    materias = MATERIAS_VIGENTE[y] or []
    add_group_row(ws_v, row, f"  {YEAR_NAMES[y-1]}"); row += 1
    for s, nombre in enumerate(materias):
        abrev = ABREV_MAP.get(nombre, 'XX')
        cal_rows = [
            (f"calif.{y}.{s}.materia", f"{s+1}. {nombre}", f"DisplayData.calificaciones[{YEAR_NAMES[y-1]}][{s}].materia", "", f"calificaciones[{YEAR_NAMES[y-1]}][{s}].materia", ""),
            (f"calif.{y}.{s}.nota", f"{s+1}. {nombre} - Nota", f"DisplayData.calificaciones[{YEAR_NAMES[y-1]}][{s}].nota", "", f"calificaciones[{YEAR_NAMES[y-1]}][{s}].nota", ""),
            (f"calif.{y}.{s}.literal", f"{s+1}. {nombre} - Literal", f"DisplayData.calificaciones[{YEAR_NAMES[y-1]}][{s}].literal", "", f"calificaciones[{YEAR_NAMES[y-1]}][{s}].literal", ""),
            (f"calif.{y}.{s}.te", f"{s+1}. {nombre} - T-E", f"DisplayData.calificaciones[{YEAR_NAMES[y-1]}][{s}].tipoEvaluacion", "", f"calificaciones[{YEAR_NAMES[y-1]}][{s}].tipoEvaluacion", ""),
            (f"calif.{y}.{s}.mes", f"{s+1}. {nombre} - Mes", f"DisplayData.calificaciones[{YEAR_NAMES[y-1]}][{s}].fechaMes", "", f"calificaciones[{YEAR_NAMES[y-1]}][{s}].fechaMes", ""),
            (f"calif.{y}.{s}.anio", f"{s+1}. {nombre} - Anio", f"DisplayData.calificaciones[{YEAR_NAMES[y-1]}][{s}].fechaAnio", "", f"calificaciones[{YEAR_NAMES[y-1]}][{s}].fechaAnio", ""),
            (f"calif.{y}.{s}.inst", f"{s+1}. {nombre} - Inst. Educ.", f"DisplayData.calificaciones[{YEAR_NAMES[y-1]}][{s}].instEduc", "", f"calificaciones[{YEAR_NAMES[y-1]}][{s}].instEduc", ""),
            (f"calif.{y}.{s}.numero", f"{s+1}. {nombre} - Nro.", f"DisplayData.calificaciones[{YEAR_NAMES[y-1]}][{s}].numero", "", f"calificaciones[{YEAR_NAMES[y-1]}][{s}].numero", ""),
        ]
        for b, l, f, rk, df, n in cal_rows:
            add_data_row(ws_v, row, b, l, f, rk, df, n, False); row += 1

# ----- CALIFICACIONES (rawData) -----
add_group_row(ws_v, row, "CALIFICACIONES (rawData: rawData.CAMPO.ABREV.ANIO)"); row += 1

# rawData fields for calificaciones
MATERIAS_RAW = ['CA-Castellano', 'IN-Ingles', 'MA-Matematicas', 'EF-Educacion Fisica',
                'AP-Arte y Patrimonio', 'CN-Ciencias Naturales', 'GH-Geog. Hist. y Ciudad.',
                'FI-Fisica', 'QU-Quimica', 'BI-Biologia', 'CT-Ciencias de la Tierra',
                'FSN-Form. Soberania Nacional']
RAW_FIELDS = ['NOTA', 'LITERAL', 'EVAL', 'MES', 'ANIO', 'INST']

for y in range(1, 6):
    add_group_row(ws_v, row, f"  {YEAR_NAMES[y-1]} (rawData)"); row += 1
    materias = MATERIAS_VIGENTE[y] or []
    for nombre in materias:
        abrev = ABREV_MAP.get(nombre, 'XX')
        for field in RAW_FIELDS:
            binding = f"rawData.{field}.{abrev}.{y}"
            label = f"{YEAR_NAMES[y-1]} - {nombre} - {field.capitalize()}"
            raw_key = f"{field}.{abrev}.{y}"
            add_data_row(ws_v, row, binding, label, "rawDataMap", raw_key, "", "", True); row += 1

# ----- ORIENTACION -----
add_group_row(ws_v, row, "ORIENTACION Y CONVIVENCIA"); row += 1
for i in range(5):
    orient_rows = [
        (f"orient.{i}.anio", f"{i+1}er Anio - Anio Escolar", f"DisplayData.orientacion[{i}].anio", "", f"orientacion[{i}].anio", ""),
        (f"orient.{i}.literal", f"{i+1}er Anio - Literal", f"DisplayData.orientacion[{i}].literal", "", f"orientacion[{i}].literal", ""),
    ]
    for b, l, f, rk, df, n in orient_rows:
        add_data_row(ws_v, row, b, l, f, rk, df, n, False); row += 1

# ----- GRUPOS -----
add_group_row(ws_v, row, "GRUPOS (CREACION/RECREACION)"); row += 1
for i in range(5):
    grupo_rows = [
        (f"grupo.{i}.anio", f"{i+1}er Anio - Anio Escolar", f"DisplayData.grupos[{i}].anio", "", f"grupos[{i}].anio", ""),
        (f"grupo.{i}.grupo", f"{i+1}er Anio - Nombre Grupo", f"DisplayData.grupos[{i}].grupo", "", f"grupos[{i}].grupo", ""),
        (f"grupo.{i}.literal", f"{i+1}er Anio - Literal", f"DisplayData.grupos[{i}].literal", "", f"grupos[{i}].literal", ""),
    ]
    for b, l, f, rk, df, n in grupo_rows:
        add_data_row(ws_v, row, b, l, f, rk, df, n, False); row += 1

# ----- PLANTEL POR MATERIA (rawData) -----
add_group_row(ws_v, row, "PLANTEL POR MATERIA (rawData - Constancia)"); row += 1
PLT_FIELDS = ['INST_NAME', 'INST_LOCAL', 'INST_EF']
PLT_LABELS = {'INST_NAME': 'Nombre Plantel', 'INST_LOCAL': 'Localidad Plantel', 'INST_EF': 'E.F. Plantel'}

for y in range(1, 6):
    add_group_row(ws_v, row, f"  Anio {y}"); row += 1
    materias = MATERIAS_VIGENTE[y] or []
    for nombre in materias:
        abrev = ABREV_MAP.get(nombre, 'XX')
        for pf in PLT_FIELDS:
            binding = f"rawData.{pf}.{abrev}.{y}"
            label = f"{PLT_LABELS[pf]} - {nombre} ({y})"
            raw_key = f"{pf}.{abrev}.{y}"
            add_data_row(ws_v, row, binding, label, "rawDataMap", raw_key, "", f"Se resuelve desde instituciones usando slot numbers", True); row += 1

# ----- SECCIONES (rawData) -----
add_group_row(ws_v, row, "SECCIONES (rawData)"); row += 1
for i in range(1, 6):
    binding = f"rawData.SECCION.{i}"
    label = f"Seccion {i}"
    raw_key = f"SECCION.{i}"
    add_data_row(ws_v, row, binding, label, "rawDataMap", raw_key, "", "", True); row += 1

# ----- PROMEDIOS (rawData) -----
add_group_row(ws_v, row, "PROMEDIOS (rawData)"); row += 1
prom_rows = [
    ("rawData.PROMEDIO.BASICA", "Promedio Basica", "rawDataMap", "PROMEDIO.BASICA", "", ""),
    ("rawData.PROMEDIO.DIVERSIFICADO", "Promedio Diversificado", "rawDataMap", "PROMEDIO.DIVERSIFICADO", "", ""),
    ("rawData.PROMEDIO.TOTAL", "Promedio Total", "rawDataMap", "PROMEDIO.TOTAL", "", ""),
]
for b, l, f, rk, df, n in prom_rows:
    add_data_row(ws_v, row, b, l, f, rk, df, n, True); row += 1

# Freeze panes
ws_v.freeze_panes = "A2"
ws_v.auto_filter.ref = f"A1:G{row-1}"


# =============================================
# HOJA 2: PLAN DEROGADO
# =============================================
ws_d = wb.create_sheet("Plan Derogado")

# Headers Derogado
for col, h in enumerate(headers_v, 1):
    cell = ws_d.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill(start_color="7B2D26", end_color="7B2D26", fill_type="solid")
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    cell.border = THIN_BORDER

for i, w in enumerate(widths_v, 1):
    ws_d.column_dimensions[get_column_letter(i)].width = w
ws_d.row_dimensions[1].height = 30

row_d = 2
GROUP_FILL_D = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
GROUP_FONT_D = Font(name="Calibri", size=11, bold=True, color="7B2D26")

def add_group_row_d(ws, r, group_name):
    for col in range(1, 8):
        c = ws.cell(row=r, column=col)
        c.fill = GROUP_FILL_D
        c.border = THIN_BORDER
    ws.cell(row=r, column=1, value=group_name).font = GROUP_FONT_D

# ----- DATOS PERSONALES -----
add_group_row_d(ws_d, row_d, "DATOS PERSONALES (rawData)"); row_d += 1
pers_d = [
    ("rawData.CEDULA", "Cedula", "rawDataMap", "CEDULA", "", "Dato personal directo"),
    ("rawData.FECHA", "Fecha", "rawDataMap", "FECHA", "", "Fecha de nacimiento"),
    ("rawData.APELLIDOS", "Apellidos", "rawDataMap", "APELLIDOS", "", ""),
    ("rawData.NOMBRES", "Nombres", "rawDataMap", "NOMBRES", "", ""),
    ("rawData.PAIS", "Pais", "rawDataMap", "PAIS", "", ""),
    ("rawData.ESTADO", "Estado", "rawDataMap", "ESTADO", "", ""),
    ("rawData.MUNICIPIO", "Municipio", "rawDataMap", "MUNICIPIO", "", ""),
    ("rawData.LUGAR", "Lugar", "rawDataMap", "LUGAR", "", ""),
    ("rawData.PROMEDIO.BASICA", "Promedio Academico Basica", "rawDataMap", "PROMEDIO.BASICA", "", ""),
    ("rawData.PROMEDIO.DIVERSIFICADO", "Promedio Academico Diversificado", "rawDataMap", "PROMEDIO.DIVERSIFICADO", "", ""),
    ("expedicion.fecha", "Fecha de Expedicion", "rawDataMap.EXPEDICION.FECHA o DisplayData.fechaExpedicion", "EXPEDICION.FECHA", "fechaExpedicion", "Dashboard"),
    ("expedicion.lugar", "Lugar de Expedicion", "rawDataMap.EXPEDICION.LUGAR o DisplayData.lugar", "EXPEDICION.LUGAR", "lugar", "Dashboard"),
    ("director.nombre", "Nombre del Director", "rawDataMap.DIRECTOR.NOMBRE o DisplayData.director", "DIRECTOR.NOMBRE", "director.apellidosNombres", "Dashboard"),
    ("director.cedula", "Cedula del Director", "rawDataMap.DIRECTOR.CEDULA o DisplayData.director", "DIRECTOR.CEDULA", "director.cedula", "Dashboard"),
]
for b, l, f, rk, df, n in pers_d:
    is_raw = b.startswith("rawData.")
    add_data_row(ws_d, row_d, b, l, f, rk, df, n, is_raw); row_d += 1

# ----- DOCUMENTO -----
add_group_row_d(ws_d, row_d, "DOCUMENTO (rawData)"); row_d += 1
doc_d = [
    ("rawData.SERIALTITULO", "Serial T.", "rawDataMap", "SERIALTITULO", "", ""),
    ("rawData.FECHAEMISIONT", "Fecha Emision T.", "rawDataMap", "FECHAEMISIONT", "", ""),
    ("rawData.EGRESOAÑO", "Anio Egreso T.", "rawDataMap", "EGRESOAÑO", "", ""),
    ("rawData.FECHAEMISIONN", "Fecha Emision N.", "rawDataMap", "FECHAEMISIONN", "", ""),
    ("rawData.PROMEDIO.TOTAL", "Promedio Total", "rawDataMap", "PROMEDIO.TOTAL", "", ""),
    ("rawData.ACTA", "Acta", "rawDataMap", "ACTA", "", ""),
]
for b, l, f, rk, df, n in doc_d:
    add_data_row(ws_d, row_d, b, l, f, rk, df, n, True); row_d += 1

# ----- EDUCACION BASICA -----
add_group_row_d(ws_d, row_d, "EDUCACION BASICA (rawData)"); row_d += 1
for i in range(1, 6):
    bas_d = [
        (f"rawData.INST.BASICA.{i}", f"Institucion Basica {i}", "rawDataMap", f"INST.BASICA.{i}", "", ""),
        (f"rawData.LOCAL.BASICA.{i}", f"Localidad Basica {i}", "rawDataMap", f"LOCAL.BASICA.{i}", "", ""),
        (f"rawData.EF.BASICA.{i}", f"E.F. Basica {i}", "rawDataMap", f"EF.BASICA.{i}", "", ""),
    ]
    for b, l, f, rk, df, n in bas_d:
        add_data_row(ws_d, row_d, b, l, f, rk, df, n, True); row_d += 1

# Obs Basica
for i in range(1, 6):
    binding = f"rawData.OBS.BASICA.L{i}"
    label = f"Obs. Basica - Linea {i}"
    add_data_row(ws_d, row_d, binding, label, "rawDataMap", f"OBS.BASICA.L{i}", "", "", True); row_d += 1

# ----- DIVERSIFICADO -----
add_group_row_d(ws_d, row_d, "DIVERSIFICADO (rawData)"); row_d += 1
for i in range(1, 6):
    div_d = [
        (f"rawData.INST.DIV.{i}", f"Institucion Diversificado {i}", "rawDataMap", f"INST.DIV.{i}", "", ""),
        (f"rawData.LOCAL.DIV.{i}", f"Localidad Diversificado {i}", "rawDataMap", f"LOCAL.DIV.{i}", "", ""),
        (f"rawData.EF.DIV.{i}", f"E.F. Diversificado {i}", "rawDataMap", f"EF.DIV.{i}", "", ""),
    ]
    for b, l, f, rk, df, n in div_d:
        add_data_row(ws_d, row_d, b, l, f, rk, df, n, True); row_d += 1

# Obs Div
for i in range(1, 6):
    binding = f"rawData.OBS.DIV.L{i}"
    label = f"Obs. Diversificado - Linea {i}"
    add_data_row(ws_d, row_d, binding, label, "rawDataMap", f"OBS.DIV.L{i}", "", "", True); row_d += 1

# ----- CALIFICACIONES DEROGADO (rawData) -----
MATERIAS_DEROGADO_RAW = [
    ('CA', 'Castellano'), ('IN', 'Ingles'), ('MA', 'Matematicas'),
    ('EN', 'Est. de la Naturaleza'), ('HV', 'Historia de Venezuela'),
    ('EF', 'Educacion Fisica'),
    ('EFC', 'Educacion para el Trabajo'), ('GG', 'Guidancia y Bienestar'),
    ('EA', 'Educacion Artistica'),
    ('EPT', 'Ed. Physica y Tecnologia'),
    ('EPS', 'Ed. para la Salud'),
    ('CB', 'Ciencias Biologicas'), ('HU', 'Humanidades'),
    ('ET', 'Ed. Tecnologica'),
    ('HC', 'Historia y Ciencias Sociales'),
    ('DT', 'Dibujo Tecnico'), ('FIL', 'Filosofia'),
    ('IPM', 'Instruccion Premilitar'),
    ('GEV', 'Geografia, Economia y Venezolanidad'),
    ('CT', 'Ciencias de la Tierra'), ('QU', 'Quimica'),
    ('FI', 'Fisica'), ('BI', 'Biologia'),
]

add_group_row_d(ws_d, row_d, "CALIFICACIONES (rawData - Derogado)"); row_d += 1

for y in range(1, 6):
    add_group_row_d(ws_d, row_d, f"  {YEAR_NAMES[y-1]}"); row_d += 1
    for abrev, nombre in MATERIAS_DEROGADO_RAW:
        for field in RAW_FIELDS:
            binding = f"rawData.{field}.{abrev}.{y}"
            label = f"{YEAR_NAMES[y-1]} - {nombre} - {field.capitalize()}"
            raw_key = f"{field}.{abrev}.{y}"
            add_data_row(ws_d, row_d, binding, label, "rawDataMap", raw_key, "", "", True); row_d += 1

# ----- EPT (Derogado only) -----
add_group_row_d(ws_d, row_d, "EPT - EDUCACION FISICA Y TECNOLOGIA (Derogado)"); row_d += 1
for i in range(1, 13):
    ept_rows = [
        (f"rawData.EPT.GRADO.{i}", f"EPT - Grado {i}", "rawDataMap", f"EPT.GRADO.{i}", "", ""),
        (f"rawData.EPT.NOMBRE.{i}", f"EPT - Nombre {i}", "rawDataMap", f"EPT.NOMBRE.{i}", "", ""),
        (f"rawData.EPT.HORAS.{i}", f"EPT - Horas {i}", "rawDataMap", f"EPT.HORAS.{i}", "", ""),
    ]
    for b, l, f, rk, df, n in ept_rows:
        add_data_row(ws_d, row_d, b, l, f, rk, df, n, True); row_d += 1

# ----- LITERALES FINALES (rawData) -----
add_group_row_d(ws_d, row_d, "LITERALES FINALES (rawData)"); row_d += 1
for i in range(1, 6):
    binding = f"rawData.LITERAL.FINAL.{i}"
    label = f"Literal Final - Anio {i}"
    add_data_row(ws_d, row_d, binding, label, "rawDataMap", f"LITERAL.FINAL.{i}", "", "", True); row_d += 1

# ----- SECCIONES -----
add_group_row_d(ws_d, row_d, "SECCIONES (rawData)"); row_d += 1
for i in range(1, 6):
    binding = f"rawData.SECCION.{i}"
    label = f"Seccion {i}"
    add_data_row(ws_d, row_d, binding, label, "rawDataMap", f"SECCION.{i}", "", "", True); row_d += 1

# Freeze panes
ws_d.freeze_panes = "A2"
ws_d.auto_filter.ref = f"A1:G{row_d-1}"


# =============================================
# HOJA 3: RESUMEN / LEGENDO
# =============================================
ws_r = wb.create_sheet("Leyenda y Resumen")

ws_r.column_dimensions['A'].width = 20
ws_r.column_dimensions['B'].width = 80

r = 1
ws_r.cell(row=r, column=1, value="Campo").font = HEADER_FONT
ws_r.cell(row=r, column=1).fill = HEADER_FILL
ws_r.cell(row=r, column=1).border = THIN_BORDER
ws_r.cell(row=r, column=2, value="Descripcion").font = HEADER_FONT
ws_r.cell(row=r, column=2).fill = HEADER_FILL
ws_r.cell(row=r, column=2).border = THIN_BORDER

legend = [
    ("Binding (dataBinding)", "El string que se coloca en la celda del Editor de Formatos. Es el path que usa resolveBinding() para buscar el dato."),
    ("Descripcion (Editor)", "El texto que se ve en el combobox del editor de formatos al seleccionar un binding."),
    ("Fuente de Datos", "Indica de donde viene el dato: 'DisplayData' (datos estructurados del API) o 'rawDataMap' (mapa plano de rawData)."),
    ("Clave rawDataMap", "La clave exacta en el diccionario rawDataMap. Solo aplica para bindings que empiezan con 'rawData.'"),
    ("Campo en DisplayData", "El campo exacto dentro del objeto DisplayData. Solo aplica para bindings estructurados (student., school., doc., etc.)"),
    ("Notas", "Informacion adicional sobre ese binding en particular."),
    ("", ""),
    ("COLOR VERDE (Consolas)", "Binding de tipo rawData - busca directamente en rawDataMap[key]"),
    ("COLOR ROJO (Consolas)", "Binding estructurado - resuelve via DisplayData (student., school., doc., etc.)"),
    ("", ""),
    ("FUNCIONAMIENTO", ""),
    ("resolveBinding(path, data)", "Funcion principal. Recibe el string del binding y el DisplayData, devuelve el texto a mostrar."),
    ("", "Si el path incluye {{ }} (template literal), reemplaza cada {{binding}} internamente."),
    ("", "Si el path incluye comas, separa y concatena (soporta textos literales entre comillas)."),
    ("", "Si no, resuelve como binding simple."),
    ("", ""),
    ("Dashboard Override", "Los datos del Dashboard (celdas [3][25], [3][33], [5][25], [6][25]) se sobreescriben en rawDataMap."),
    ("", "Esto permite que el usuario cambie fecha, lugar, director directamente en el dashboard."),
    ("", ""),
    ("PLAN VIGENTE vs DEROGADO", ""),
    ("Plan Vigente", "Usa bindings estructurados (student., calif., orient., etc.) y rawData.* para campos planos."),
    ("Plan Derogado", "Usa casi exclusivamente rawData.* porque los datos vienen en formato crudo/plano."),
    ("", ""),
    ("SYNTAX ESPECIAL", ""),
    ('Template literal', 'Ej: "Serial: {{doc.acta}} - Fecha: {{doc.actaFecha}}" - mezcla texto fijo con bindings.'),
    ("Comma-separated", 'Ej: "doc.acta, \" - \", doc.actaFecha" - concatena bindings con separadores.'),
    ("dateFormat", "En CellConfig, formatea la fecha resuelta: 'DD/MM/YYYY', 'DD DE MES DE YYYY', etc."),
]

for item in legend:
    r += 1
    ws_r.cell(row=r, column=1, value=item[0]).font = Font(name="Calibri", size=10, bold=True)
    ws_r.cell(row=r, column=1).border = THIN_BORDER
    ws_r.cell(row=r, column=2, value=item[1]).font = DATA_FONT
    ws_r.cell(row=r, column=2).border = THIN_BORDER
    ws_r.cell(row=r, column=2).alignment = WRAP

# Save
output_path = "/home/z/my-project/download/Bindings_jo-sigae.xlsx"
wb.save(output_path)
print(f"Excel generado exitosamente: {output_path}")
print(f"Hoja Vigente: filas de datos = {row - 2}")
print(f"Hoja Derogado: filas de datos = {row_d - 2}")
