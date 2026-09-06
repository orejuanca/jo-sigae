# -*- coding: utf-8 -*-
"""
GENERADOR DE LA REJILLA EMG (v12) — lee las hojas RF 1°..5° del xlsm VERBATIM y emite
un JSON por grado con TODO el diseño: columnas/filas px, TODAS las líneas (segmentos),
merges, textos estáticos con estilo, ANCLAS DINÁMICAS en px y el logo.

REGLAS DEL USUARIO APLICADAS:
  1. ORTOGRAFÍA CORRECTA primero (las erratas del original NO se replican)
  2. Columna CÉDULA reducida: B,C,D 2.0u->1.4u y E..M 0.855u->0.55u
     (B..M: 138px -> 111px; 'V12345678900' ~73px => aire ~38px, holgado)

Geometría: px = round(u * 5.5 + 5) por columna (MDW 5.5; suma A..BN ~1100px y a escala
Excel ~74% llena el papel OFICIO 8.5in = 816px, validado en v10/v11).
Salida: jo-sigae/src/lib/rf/rejilla-{1..5}.json
"""
import json, os, re, unicodedata
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSM = 'upload/Control de Alumnos UENCC 2021 - 2022 - copia.xlsm'
HOJAS = {1: 'RF 1°', 2: 'RF 2°', 3: 'RF 3°', 4: 'RF 4°', 5: 'RF 5°'}
DEST = 'jo-sigae/src/lib/rf'
MDW = 5.5

SUBSTR = [
    ('PARTICIPACIÓN EN GRUPOS DECREACIÓN, RECREACIÓN Y PRODUCCIÓN',
     'PARTICIPACIÓN EN GRUPOS DE CREACIÓN, RECREACIÓN Y PRODUCCIÓN'),
    ('Cédula de \nIdentidad', 'Cédula de\nIdentidad'),
    ('FECHA DE \nNACIMIENTO', 'FECHA DE\nNACIMIENTO'),
    ('SELLO DE LA ZONA \nEDUCATIVA', 'SELLO DE LA ZONA\nEDUCATIVA'),
    ('Código de la institución Educativa:', 'Código de la Institución Educativa:'),
    ('Cédula de Identidad.:', 'Cédula de Identidad:'),
]
EXACT = {
    'FIRMA': 'Firma:',
    'Firma': 'Firma:',
    'Matemática': 'Matemáticas',
    'Ciencias de La Tierra': 'Ciencias de la Tierra',
    'Apellidos y Nombres': 'Apellidos y Nombres del Profesor',   # solo col U (21), bloque V
}
EXCEP_NORM = {'Apellidos y Nombres:', 'SELLO DEL PLANTEL'}  # no tocar

def sin_acentos(s):
    return ''.join(ch for ch in unicodedata.normalize('NFD', s) if unicodedata.category(ch) != 'Mn')

def normalizar(t, r=0, c=0):
    if not isinstance(t, str):
        return t
    if t in EXCEP_NORM or t.rstrip() == 'Apellidos y Nombres:':
        return t
    for a, b in SUBSTR:
        if a in t:
            t = t.replace(a, b)
    ts = t.rstrip()
    if ts in EXACT:
        if not (ts == 'Apellidos y Nombres' and c != 21):
            t = EXACT[ts]
    t = re.sub(r' +\n', '\n', t)   # espacio de más antes del salto de línea
    return t.rstrip()

ANCHOS_OVERRIDE = {2: 1.4, 3: 1.4, 4: 1.4}
ANCHOS_OVERRIDE.update({c: 0.55 for c in range(5, 14)})

wb_v = load_workbook(XLSM, data_only=True)
wb_f = load_workbook(XLSM, data_only=False)

def col_px(u): return int(round(u * MDW + 5))
def row_px(pt): return int(round(pt * 4.0 / 3.0))

def borde(celda, lado):
    if celda is None: return False
    s = getattr(celda.border, lado)
    return bool(s and s.style)

def main():
    os.makedirs(DEST, exist_ok=True)
    for grado, hoja in HOJAS.items():
        wv, wf = wb_v[hoja], wb_f[hoja]
        # print_area no la carga openpyxl: se DERIVA del marco de bordes. Límite real =
        # última columna/fila donde el borde corre en MUCHAS filas/cols (el contorno);
        # los artefactos de helpers (BP..BT) solo tocan 1-2 filas.
        gpre = wf._cells
        filas_por_col, cols_por_fila = {}, {}
        for (r, c), cell in gpre.items():
            if r > 82 or c > 72: continue
            if any(borde(cell, l) for l in ('top', 'bottom', 'left', 'right')):
                filas_por_col[c] = filas_por_col.get(c, 0) + 1
                cols_por_fila[r] = cols_por_fila.get(r, 0) + 1
        MAXC = max((c for c, n in filas_por_col.items() if n >= 5), default=66)
        MAXR = max((r for r, n in cols_por_fila.items() if n >= 5), default=77)
        print('  marco -> MAXC=%d MAXR=%d (bordes sueltos hasta c=%d r=%d descartados)' % (
            MAXC, MAXR, max(filas_por_col), max(cols_por_fila)))

        # anchos: expandir grupos min..max de column_dimensions (openpyxl agrupa)
        wmap = {}
        for key, dim in wf.column_dimensions.items():
            if dim.width is None: continue
            mn, mx = dim.min, dim.max
            if not mn:
                mn = mx = 0
                for ch in key: mn = mn * 26 + (ord(ch) - 64)
                mx = mn
            for c in range(mn, mx + 1): wmap[c] = dim.width
        colW = []
        for c in range(1, MAXC + 1):
            u = ANCHOS_OVERRIDE.get(c) or wmap.get(c) or 8.43
            colW.append(col_px(u))
        dRH = wf.sheet_format.defaultRowHeight or 12.75
        rowH = []
        for r in range(1, MAXR + 1):
            dim = wf.row_dimensions.get(r)
            rowH.append(row_px(dim.height if dim and dim.height else dRH))
        X = [0]
        for w in colW: X.append(X[-1] + w)
        Y = [0]
        for h in rowH: Y.append(Y[-1] + h)
        W, H = X[-1], Y[-1]

        # merges
        cover = {}
        for mr in wf.merged_cells.ranges:
            if mr.max_row > MAXR or mr.max_col > MAXC: continue
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    cover[(r, c)] = (mr.min_row, mr.min_col, mr.max_row, mr.max_col)

        def rect_de(r, c):
            r1, c1, r2, c2 = cover.get((r, c), (r, c, r, c))
            return {'x': X[c1 - 1], 'y': Y[r1 - 1], 'w': X[c2] - X[c1 - 1], 'h': Y[r2] - Y[r1 - 1]}

        def estilo_de(r, c):
            est = wf._cells.get((r, c))
            f = est.font if est else None
            a = est.alignment if est else None
            ah = (a.horizontal if a else None) or 'left'
            if ah == 'general': ah = 'left'
            av = (a.vertical if a else None) or 'bottom'
            return {'pt': float(f.size if f and f.size else 10), 'b': bool(f.bold if f else False),
                    'rot': (a.textRotation if a else 0) or 0,
                    'wrap': bool(a.wrapText if a else False), 'ah': ah, 'av': av,
                    'u': bool(f.underline if f else False) and f.underline != 'none'}

        # --- líneas ---
        Hs, Vs = {}, {}
        g = wf._cells
        for r in range(1, MAXR + 2):
            for c in range(1, MAXC + 1):
                if r <= MAXR and borde(g.get((r, c)), 'top'): Hs[(r, c)] = 1
                elif r > 1 and borde(g.get((r - 1, c)), 'bottom'): Hs[(r, c)] = 1
        for r in range(1, MAXR + 1):
            for c in range(1, MAXC + 2):
                if c <= MAXC and borde(g.get((r, c)), 'left'): Vs[(r, c)] = 1
                elif c > 1 and borde(g.get((r, c - 1)), 'right'): Vs[(r, c)] = 1

        # --- v13: LINEAS FANTASMA. Excel NO dibuja bordes internos de celdas
        # combinadas: un segmento H entre (r-1,c) y (r,c) es fantasma si ambas
        # celdas pertenecen al MISMO merge (los del perimetro del merge SI se
        # dibujan y se conservan). Idem V entre (r,c-1) y (r,c).
        def en_mismo_merge(r1, c1, r2, c2):
            m1 = cover.get((r1, c1), (r1, c1, r1, c1))
            m2 = cover.get((r2, c2), (r2, c2, r2, c2))
            return m1 == m2
        Hs = {k: v for k, v in Hs.items()
              if not (k[0] > 1 and en_mismo_merge(k[0] - 1, k[1], k[0], k[1]))}
        Vs = {k: v for k, v in Vs.items()
              if not (k[1] > 1 and en_mismo_merge(k[0], k[1] - 1, k[0], k[1]))}

        def consolidar(d, es_h):
            out = []
            if es_h:
                por = {}
                for (r, c) in d: por.setdefault(r, []).append(c)
                for r, cols in por.items():
                    cols.sort(); a = b = cols[0]
                    for n in cols[1:] + [None]:
                        if n == b + 1: b = n; continue
                        out.append([Y[r - 1], X[a - 1], X[b]])
                        if n: a = b = n
            else:
                por = {}
                for (r, c) in d: por.setdefault(c, []).append(r)
                for c, rows in por.items():
                    rows.sort(); a = b = rows[0]
                    for n in rows[1:] + [None]:
                        if n == b + 1: b = n; continue
                        out.append([X[c - 1], Y[a - 1], Y[b]])
                        if n: a = b = n
            out.sort()
            return out
        lineas = {'H': consolidar(Hs, True), 'V': consolidar(Vs, False)}

        # ---- localización de zonas dinámicas ----
        def ubicar(t, f1=1, f2=None, c1=1, c2=None):
            f2 = f2 or MAXR; c2 = c2 or MAXC; t = t.strip()
            for r in range(f1, f2 + 1):
                for c in range(c1, c2 + 1):
                    v = wv._cells.get((r, c))
                    if v is not None and isinstance(v.value, str) and v.value.strip() == t:
                        return r, c
            for r in range(f1, f2 + 1):
                for c in range(c1, c2 + 1):
                    v = wv._cells.get((r, c))
                    if v is not None and isinstance(v.value, str) and t in v.value:
                        return r, c
            return None, None

        def anchor(r, c):
            p = rect_de(r, c)
            e = estilo_de(r, c)
            return {'x': p['x'], 'y': p['y'], 'w': p['w'], 'h': p['h'],
                    'ah': e['ah'], 'av': e['av'], 'pt': e['pt'], 'b': e['b']}

        zonas = set()
        def marcar(r, c, clave):
            if not r: return
            zonas.add((r, c))
            DIN[clave] = anchor(r, c)
        DIN = {}

        rr, cc = ubicar('2021 - 2022', 1, 5); marcar(rr, cc, 'ano')
        TIPOS = ['NO CURSANTE', 'PENDIENTE', 'EQUIVALENCIA', 'QUEDADA', 'TRANSFERENCIA', 'FINAL', 'REVISIÓN']
        MESES = ['SEPTIEMBRE - 2021', 'OCTUBRE - 2021', 'NOVIEMBRE - 2021', 'DICIEMBRE - 2021',
                 'ENERO - 2022', 'FEBRERO - 2022', 'MARZO - 2022', 'ABRIL - 2022', 'MAYO - 2022',
                 'JUNIO - 2022', 'JULIO - 2022', 'AGOSTO - 2022']
        rrt = cct = None
        # v13: pasada EXACTA en filas 1..5 (la pasada 'contiene' de ubicar caia en
        # 'RESUMEN FINAL DEL RENDIMIENTO ESTUDIANTIL' cuando el tipo cacheado no era
        # FINAL -> en RF 3° el anchor tipo apuntaba al TITULO y este no salia)
        # comparacion SIN acentos ('REVISION' cacheado en 3° == 'REVISIÓN')
        TIPOS_U = {sin_acentos(t).upper(): t for t in TIPOS}
        for r in range(1, 6):
            for c in range(1, MAXC + 1):
                v = wv._cells.get((r, c))
                if v is not None and isinstance(v.value, str) and sin_acentos(v.value.strip()).upper() in TIPOS_U:
                    rrt, cct = r, c; break
            if rrt: break
        if not rrt:
            re2, ce2 = ubicar('Tipo de Evaluación:', 1, 5)
            if re2: rrt, cct = re2, ce2 + 2
        marcar(rrt, cct, 'tipo')
        rrm = ccm = None
        for t in MESES:
            rrm, ccm = ubicar(t, 1, 5)
            if rrm: break
        if not rrm:
            re2, ce2 = ubicar('Mes y Año:', 1, 5)
            if re2: rrm, ccm = re2, ce2 + 2
        marcar(rrm, ccm, 'mes')
        rr, cc = ubicar('OD16751520', 5, 10); marcar(rr, cc, 'cod')
        rr, cc = ubicar('U E N CREACION CUA', 5, 10); marcar(rr, cc, 'denom')
        rr, cc = ubicar('URB. JOSE DE SAN MARTIN', 5, 10); marcar(rr, cc, 'dir')
        rr, cc = ubicar('(0239) 7163530', 5, 10); marcar(rr, cc, 'tel')
        rr, cc = ubicar('RAFAEL URDANETA', 5, 10); marcar(rr, cc, 'munic')
        hits = []
        for r in range(5, 10):
            for c in range(1, MAXC + 1):
                v = wv._cells.get((r, c))
                if v is not None and v.value == 'MIRANDA': hits.append((r, c))
        if len(hits) >= 2:
            marcar(hits[0][0], hits[0][1], 'entFed'); marcar(hits[1][0], hits[1][1], 'cdcee')
        rr, cc = ubicar('PAREDES HURTADO, RAQUEL', 5, 10); marcar(rr, cc, 'director')
        rr, cc = ubicar('V 6419439', 5, 10); marcar(rr, cc, 'cedDir')
        # VII
        rr, cc = ubicar('VII. Observaciones:', MAXR - 12, MAXR)
        obs1 = None
        if rr:
            for c in (13, 14, 15, 16):
                if (rr, c) in cover and cover[(rr, c)][1] == c:
                    obs1 = c; break
            marcar(rr, obs1 or 14, 'obs1')
            marcar(rr + 1, 1, 'obs2')
        # VI
        rr, cc = ubicar('SECCIÓN', 55, MAXR, 55, MAXC)
        if rr: marcar(rr + 1, cc, 'seccion')
        rr, cc = ubicar('N° DE ESTUDIANTES POR SECCIÓN', 55, MAXR, 55, MAXC)
        if rr:
            nr = None
            for dr in (1, 2, 3, 4):
                v = wv._cells.get((rr + dr, cc))
                if v is not None and isinstance(v.value, (int, float)):
                    marcar(rr + dr, cc, 'nSeccion'); nr = rr + dr; break
            if nr is None:
                marcar(rr + 1, cc, 'nSeccion'); nr = rr + 1
            hecho = False
            for c2 in range(cc + 3, MAXC + 1):
                v = wv._cells.get((nr, c2))
                if v is not None and isinstance(v.value, (int, float)):
                    marcar(nr, c2, 'nPagina'); hecho = True; break
        # VIII (director/cédula de abajo)
        for r in range(MAXR - 10, MAXR + 1):
            for c in range(1, 30):
                v = wv._cells.get((r, c))
                if v is not None and v.value == 'PAREDES HURTADO, RAQUEL': marcar(r, c, 'director2')
                if v is not None and v.value == 'V 6419439': marcar(r, c, 'cedDir2')
        # bloque V (profesores)
        rr, cc = ubicar('Apellidos y Nombres', 55, MAXR, 1, 40)
        fila_v = rr
        prof = []
        if fila_v:
            n_areas = {1: 9, 2: 9, 3: 10, 4: 11, 5: 12}[grado]
            for i in range(n_areas):
                r = fila_v + 2 + i
                for c in (21, 36, 48):
                    zonas.add((r, c))
                prof.append({'u': anchor(r, 21), 'aj': anchor(r, 36), 'av': anchor(r, 48), 'fila': i})
        # columnas de notas y GRUPO (cabeceras filas 15/16)
        col_grupo = None
        for r in (15, 16):
            if col_grupo: break
            for c in range(48, MAXC + 1):
                v = wv._cells.get((r, c))
                if v is not None and isinstance(v.value, str) and v.value.strip() == 'GRUPO':
                    col_grupo = c; break
        col_nota = []
        for c in range(48, MAXC + 1):
            if c == col_grupo: break
            v = wv._cells.get((16, c))
            if v is not None and isinstance(v.value, str) and v.value.strip() and v.value.strip() != 'GRUPO':
                col_nota.append(c)
        aster_fila = None
        for r in range(17, 52):
            v = wv._cells.get((r, 2))
            if v is not None and v.value == '***': aster_fila = r
        star_pie = None
        if col_grupo:
            for r in range(52, 57):
                v = wv._cells.get((r, col_grupo))
                if v is not None and str(v.value).strip() == '*':
                    star_pie = anchor(r, col_grupo); zonas.add((r, col_grupo)); break
        pie = []
        for c in col_nota:
            for r in range(52, 57):
                zonas.add((r, c))
            pie.append({'c': c, 'filas': [anchor(r, c) for r in range(52, 57)]})
        alumnos = []
        for r in range(17, 52):
            try:
                alumnos.append({
                    'ced': anchor(r, 2), 'ape': anchor(r, 14), 'nom': anchor(r, 23),
                    'lug': anchor(r, 34), 'ef': anchor(r, 41), 'sexo': anchor(r, 42),
                    'dia': anchor(r, 43), 'mesN': anchor(r, 44), 'anio': anchor(r, 46),
                    'notas': [anchor(r, c) for c in col_nota],
                    'grupo': anchor(r, col_grupo),
                })
            except TypeError:
                print('DEBUG %s r=%d col_grupo=%s MAXC=%d len(X)=%d cover=%s' % (
                    hoja, r, col_grupo, MAXC, len(X), cover.get((r, col_grupo))))
                raise

        # ---- celdas estáticas (texto literal, normalizado) ----
        # ZONA DE ALUMNOS (filas 17..51): SOLO la col A (N°) es estática; todo lo demás
        # es data del docente / asteriscos => dinámica (nunca va al JSON estático)
        col_nota_set = set(col_nota)
        celdas = []
        for (r, c), cell in sorted(wv._cells.items()):
            if r > MAXR or c > MAXC: continue
            if (r, c) in zonas: continue
            if 17 <= r <= 51 and c != 1: continue
            v = cell.value
            if v is None or v == '' or isinstance(v, str) and v.startswith('='): continue
            if not isinstance(v, (str, int, float)): continue
            if isinstance(v, (int, float)):
                if 52 <= r <= 56 and c in col_nota_set: continue  # totales del pie (filas exactas): dinámicos (v13: deja pasar códigos como 31059 en VI)
                t = '%02d' % int(v) if (17 <= r <= 51 and c == 1) else str(int(v) if float(v).is_integer() else v)
            else:
                if 17 <= r <= 51 and c == 1:
                    t = normalizar(v, r, c)
                else:
                    t = normalizar(v, r, c)
                if not t: continue
            e = estilo_de(r, c)
            p = rect_de(r, c)
            cld = {'x': p['x'], 'y': p['y'], 'w': p['w'], 'h': p['h'], 't': t,
                   'pt': e['pt'], 'b': e['b'], 'rot': e['rot'], 'wrap': e['wrap'],
                   'ah': e['ah'], 'av': e['av']}
            if e['u']: cld['u'] = 1           # subrayado (titulo RESUMEN FINAL...)
            celdas.append(cld)

        data = {'grado': str(grado), 'K': round(816 / W, 4), 'W': W, 'H': H,
                'colW': colW, 'rowH': rowH, 'lineas': lineas,
                'logo': {'x': 5, 'y': 2, 'w': 473, 'h': 66},
                'celdas': celdas,
                'din': {'colNota': col_nota, 'colGrupo': col_grupo, 'asterFila': aster_fila,
                        'starPie': star_pie, 'pie': pie, 'prof': prof, 'alumnos': alumnos,
                        'etq': DIN}}
        path = os.path.join(DEST, 'rejilla-%d.json' % grado)
        with open(path, 'w', encoding='utf-8') as fh:
            json.dump(data, fh, ensure_ascii=False)
        print('%s: W=%d H=%d K=%.4f | H-lin=%d V-lin=%d | celdas=%d (u=%d) | notas=%s grupo=%s | aster=%s' % (
            hoja, W, H, data['K'], len(lineas['H']), len(lineas['V']), len(celdas),
            sum(1 for c in celdas if c.get('u')),
            col_nota, col_grupo, aster_fila))

if __name__ == '__main__':
    main()
