"""
Migration script: Reads PL and score data from Excel BOLETAS sheet
and updates BoletaExtra records in Neon PostgreSQL.

Excel columns (row 8 headers, data from row 9):
  B  (2)  = CEDULA
  BO (67) = PL1
  BP (68) = PL2
  BQ (69) = PL3
  BR (70) = PL4
  BS (71) = PL5
  BU (73) = CA  (Castellano)
  BV (74) = ILE (Inglés)
  BW (75) = MA  (Matemática)
  BX (76) = EF  (Ed. Física)
  BY (77) = AP  (Arte y Patrimonio)
  BZ (78) = CN  (Cs. Naturales)
  CA (79) = GHC (Geo. Hist. y Ciudadanía)
"""

import openpyxl
import psycopg2
import sys

# ── Config ──────────────────────────────────────────────────────────────────
EXCEL_PATH = "/home/z/my-project/upload/Certificacion 32001-32011-31018-31059 UENCC - copia.xlsm"
DB_URL = "postgresql://neondb_owner:npg_vDTFWj0OGL5e@ep-proud-star-ajqhfk11-pooler.c-3.us-east-2.aws.neon.tech/neondb?sslmode=require"
SHEET_NAME = "BOLETAS"
HEADER_ROW = 8
DATA_START_ROW = 9

COL_CEDULA = 2
COL_PL = [67, 68, 69, 70, 71]          # PL1-PL5
COL_SCORES = {
    "scoreCA":  73,
    "scoreILE": 74,
    "scoreMA":  75,
    "scoreEF":  76,
    "scoreAP":  77,
    "scoreCN":  78,
    "scoreGHC": 79,
}

# ── Helpers ─────────────────────────────────────────────────────────────────
def normalize_cedula(val):
    """Remove spaces and standardize cedula for matching."""
    if val is None:
        return None
    s = str(val).strip()
    # Remove all whitespace
    s = s.replace(" ", "").replace("\t", "")
    return s if s else None

def has_any_data(row_data):
    """Check if any PL or score value is non-empty/non-zero."""
    for v in row_data["pl"]:
        if v is not None and str(v).strip() not in ("", "0", "0.0"):
            return True
    for v in row_data["scores"].values():
        if v is not None and str(v).strip() not in ("", "0", "0.0"):
            return True
    return False

def to_text(val):
    """Convert value to text for DB storage; None stays None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print(f"Opening Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    print(f"Sheet '{SHEET_NAME}' — rows: {ws.max_row}, cols: {ws.max_column}")

    # 1. Read Excel data
    rows_with_data = []
    rows_total = 0
    rows_no_cedula = 0

    for r in range(DATA_START_ROW, ws.max_row + 1):
        raw_cedula = ws.cell(row=r, column=COL_CEDULA).value
        if raw_cedula is None:
            continue

        cedula = normalize_cedula(raw_cedula)
        if cedula is None:
            rows_no_cedula += 1
            continue

        rows_total += 1

        # Extract PL values
        pl = []
        for col in COL_PL:
            pl.append(to_text(ws.cell(row=r, column=col).value))

        # Extract score values
        scores = {}
        for field, col in COL_SCORES.items():
            scores[field] = to_text(ws.cell(row=r, column=col).value)

        row_data = {"cedula": cedula, "pl": pl, "scores": scores}

        if has_any_data(row_data):
            rows_with_data.append(row_data)

    print(f"\nExcel summary:")
    print(f"  Total rows with cedula: {rows_total}")
    print(f"  Rows with data to migrate: {len(rows_with_data)}")
    print(f"  Rows skipped (all empty/zero): {rows_total - len(rows_with_data)}")
    print(f"  Rows skipped (no cedula): {rows_no_cedula}")

    if not rows_with_data:
        print("\nNo data to migrate. Exiting.")
        return

    # 2. Connect to DB
    print(f"\nConnecting to PostgreSQL...")
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    # Pre-load Student id -> cedula mapping (normalized)
    print("Loading Student cedula index...")
    cur.execute('SELECT "id", "cedula" FROM "Student"')
    student_map = {}  # normalized_cedula -> student_id
    for sid, ced in cur.fetchall():
        nc = normalize_cedula(ced)
        if nc:
            student_map[nc] = sid
    print(f"  Loaded {len(student_map)} students")

    # Pre-load BoletaExtra id -> studentId mapping
    print("Loading BoletaExtra index...")
    cur.execute('SELECT "id", "studentId" FROM "BoletaExtra"')
    boleta_map = {}  # studentId -> boleta_id
    for bid, sid in cur.fetchall():
        boleta_map[sid] = bid
    print(f"  Loaded {len(boleta_map)} boleta records")

    # 3. Process updates
    matched = 0
    not_matched = 0
    updated = 0
    no_boleta = 0
    update_details = []

    for row_data in rows_with_data:
        cedula = row_data["cedula"]

        # Find student
        student_id = student_map.get(cedula)
        if not student_id:
            not_matched += 1
            print(f"  NOT FOUND in DB: cedula={cedula}")
            continue

        # Find boleta
        boleta_id = boleta_map.get(student_id)
        if not boleta_id:
            no_boleta += 1
            print(f"  NO BOLETA for student: cedula={cedula}, studentId={student_id}")
            continue

        matched += 1

        # Build SET clause only for non-empty values
        set_parts = []
        params = []
        param_idx = 1

        for i, pl_val in enumerate(row_data["pl"]):
            if pl_val is not None:
                col_name = f"pl{i+1}"
                set_parts.append(f'"{col_name}" = %s')
                params.append(pl_val)

        for field, val in row_data["scores"].items():
            if val is not None:
                set_parts.append(f'"{field}" = %s')
                params.append(val)

        if not set_parts:
            continue

        params.append(boleta_id)
        sql = f'UPDATE "BoletaExtra" SET {", ".join(set_parts)} WHERE "id" = %s'
        cur.execute(sql, params)
        updated += cur.rowcount
        update_details.append((cedula, set_parts))

    conn.commit()

    # 4. Summary
    print("\n" + "=" * 60)
    print("MIGRATION SUMMARY")
    print("=" * 60)
    print(f"  Excel rows with data:     {len(rows_with_data)}")
    print(f"  Matched students in DB:   {matched}")
    print(f"  Cedula not found in DB:   {not_matched}")
    print(f"  Student has no BoletaExtra: {no_boleta}")
    print(f"  BoletaExtra records updated: {updated}")
    print()

    if update_details:
        print("Updated records (cedula, fields):")
        for cedula, parts in update_details:
            fields = ", ".join(p.split("=")[0].strip().strip('"') for p in parts)
            print(f"  {cedula}  →  {fields}")

    cur.close()
    conn.close()
    print("\nDone.")

if __name__ == "__main__":
    main()
