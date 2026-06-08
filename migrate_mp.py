"""
Migrate Materia Pendiente data from Excel BOLETAS sheet to BoletaExtra table.
"""

import openpyxl
import psycopg2
from psycopg2 import sql

EXCEL_PATH = "/home/z/my-project/upload/Certificacion 32001-32011-31018-31059 UENCC - copia.xlsm"
DB_URL = "postgresql://neondb_owner:npg_vDTFWj0OGL5e@ep-proud-star-ajqhfk11-pooler.c-3.us-east-2.aws.neon.tech/neondb?sslmode=require"

SHEET_NAME = "BOLETAS"
HEADER_ROW = 8
DATA_START_ROW = 9

# Column indices
COL_CEDULA = 2   # B
COL_MP1 = 56     # BD
COL_MP2 = 57     # BE
COL_MP1_M1 = 58  # BF
COL_MP1_M2 = 59  # BG
COL_MP1_M3 = 60  # BH
COL_MP1_M4 = 61  # BI
COL_MP2_M1 = 62  # BJ
COL_MP2_M2 = 63  # BK
COL_MP2_M3 = 64  # BL
COL_MP2_M4 = 65  # BM


def normalize_cedula(val):
    """Strip whitespace and uppercase the cedula. Keeps spaces as-is."""
    if val is None:
        return None
    return str(val).strip().upper()


def to_str(val):
    """Convert cell value to string, None stays None."""
    if val is None:
        return None
    return str(val).strip()


def main():
    # 1. Read Excel
    print(f"Reading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # 2. Extract MP rows
    mp_rows = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cedula = ws.cell(row, COL_CEDULA).value
        mp1 = ws.cell(row, COL_MP1).value
        mp2 = ws.cell(row, COL_MP2).value

        if not mp1 and not mp2:
            continue

        norm_cedula = normalize_cedula(cedula)
        if not norm_cedula:
            continue

        record = {
            "cedula": norm_cedula,
            "materiaPendiente1": to_str(mp1),
            "materiaPendiente2": to_str(mp2),
            "mp1m1": to_str(ws.cell(row, COL_MP1_M1).value),
            "mp1m2": to_str(ws.cell(row, COL_MP1_M2).value),
            "mp1m3": to_str(ws.cell(row, COL_MP1_M3).value),
            "mp1m4": to_str(ws.cell(row, COL_MP1_M4).value),
            "mp2m1": to_str(ws.cell(row, COL_MP2_M1).value),
            "mp2m2": to_str(ws.cell(row, COL_MP2_M2).value),
            "mp2m3": to_str(ws.cell(row, COL_MP2_M3).value),
            "mp2m4": to_str(ws.cell(row, COL_MP2_M4).value),
        }
        mp_rows.append(record)

    print(f"Found {len(mp_rows)} rows with MP data in Excel.")

    if not mp_rows:
        print("No data to migrate. Exiting.")
        return

    # 3. Connect to DB
    print("Connecting to PostgreSQL...")
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    # 4. Update each record
    updated = 0
    not_found = []

    for rec in mp_rows:
        # Find studentId from Student table by cedula
        # Try exact match first, then try with/without space
        cur.execute(
            'SELECT id FROM "Student" WHERE cedula = %s OR REPLACE(cedula, \' \', \'\') = REPLACE(%s, \' \', \'\')',
            (rec["cedula"], rec["cedula"])
        )
        student_row = cur.fetchone()

        if not student_row:
            not_found.append(rec["cedula"])
            print(f"  WARNING: No student found for cedula={rec['cedula']}")
            continue

        student_id = student_row[0]

        # Check if BoletaExtra exists for this student
        cur.execute(
            'SELECT id FROM "BoletaExtra" WHERE "studentId" = %s',
            (student_id,)
        )
        boleta_row = cur.fetchone()

        if not boleta_row:
            not_found.append(rec["cedula"] + " (no BoletaExtra)")
            print(f"  WARNING: No BoletaExtra for studentId={student_id} (cedula={rec['cedula']})")
            continue

        # Update BoletaExtra
        cur.execute(
            """
            UPDATE "BoletaExtra"
            SET "materiaPendiente1" = %s,
                "materiaPendiente2" = %s,
                "mp1m1" = %s,
                "mp1m2" = %s,
                "mp1m3" = %s,
                "mp1m4" = %s,
                "mp2m1" = %s,
                "mp2m2" = %s,
                "mp2m3" = %s,
                "mp2m4" = %s,
                "updatedAt" = NOW()
            WHERE "studentId" = %s
            """,
            (
                rec["materiaPendiente1"],
                rec["materiaPendiente2"],
                rec["mp1m1"],
                rec["mp1m2"],
                rec["mp1m3"],
                rec["mp1m4"],
                rec["mp2m1"],
                rec["mp2m2"],
                rec["mp2m3"],
                rec["mp2m4"],
                student_id,
            )
        )
        updated += 1
        print(f"  Updated: cedula={rec['cedula']} (studentId={student_id})")

    conn.commit()
    cur.close()
    conn.close()

    # 5. Summary
    print("\n" + "=" * 50)
    print("MIGRATION SUMMARY")
    print("=" * 50)
    print(f"  Excel rows with MP data:  {len(mp_rows)}")
    print(f"  Successfully updated:      {updated}")
    print(f"  Not found / skipped:       {len(not_found)}")
    if not_found:
        print(f"  Missing cedulas: {not_found}")


if __name__ == "__main__":
    main()
