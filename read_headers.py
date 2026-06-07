"""
Read ALL column headers from the BOLETAS sheet of the Excel file.
Prints every column header with column letter from row 1 to row 10 to find the header row.
"""

import openpyxl
from openpyxl.utils import get_column_letter

EXCEL_PATH = "/home/z/my-project/upload/Certificacion 32001-32011-31018-31059 UENCC - copia.xlsm"
SHEET_NAME = "BOLETAS"

def main():
    print(f"Opening: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Check sheet names
    print(f"Available sheets: {wb.sheetnames}")
    
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found!")
        return
    
    ws = wb[SHEET_NAME]
    print(f"Sheet '{SHEET_NAME}' — rows: {ws.max_row}, cols: {ws.max_column}")
    print()
    
    # Print rows 1-10 to find the header row
    print("=" * 80)
    print("SCANNING ROWS 1-10 TO FIND HEADER ROW")
    print("=" * 80)
    
    for row_num in range(1, 11):
        if row_num > ws.max_row:
            break
        row_data = []
        has_content = False
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row_num, column=col).value
            if val is not None:
                has_content = True
                letter = get_column_letter(col)
                row_data.append(f"{letter}={val}")
        if has_content:
            print(f"\nRow {row_num}: {len(row_data)} non-empty cells")
            for item in row_data:
                print(f"  {item}")
        else:
            print(f"\nRow {row_num}: [EMPTY ROW]")
    
    # Based on existing scripts, header is at row 8
    # Now print ALL columns from the header row (row 8)
    print("\n" + "=" * 80)
    print("ALL COLUMN HEADERS (Row 8 - HEADER ROW per existing scripts)")
    print("=" * 80)
    
    header_row = 8
    total_cols = ws.max_column
    print(f"Total columns in sheet: {total_cols}")
    print()
    
    for col in range(1, total_cols + 1):
        letter = get_column_letter(col)
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            text = str(val).strip()
            print(f"Col {col:>3} ({letter:>3}): {text}")
        else:
            print(f"Col {col:>3} ({letter:>3}): [EMPTY]")
    
    # Also print row 9 (first data row) for comparison
    print("\n" + "=" * 80)
    print("ROW 9 (FIRST DATA ROW) — first 80 columns for reference")
    print("=" * 80)
    
    data_row = 9
    for col in range(1, min(total_cols + 1, 81)):
        letter = get_column_letter(col)
        val = ws.cell(row=data_row, column=col).value
        if val is not None:
            text = str(val).strip()
            if text:
                print(f"Col {col:>3} ({letter:>3}): {text}")
            else:
                print(f"Col {col:>3} ({letter:>3}): [EMPTY-STR]")
        else:
            print(f"Col {col:>3} ({letter:>3}): [NONE]")

if __name__ == "__main__":
    main()
